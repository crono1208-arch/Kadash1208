import tkinter as tk
from tkinter import mainloop, ttk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
import re
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

class ConciliadorApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Conciliador de Cuentas")
        
        # Configuracion de la interfaz
        self.frame = ttk.Frame(self.master)
        self.frame.pack(padx=10, pady=10)

        self.label1 = ttk.Label(self.frame, text="Archivo 1:")
        self.label1.grid(row=0, column=0, sticky=tk.W)
        self.entry1 = ttk.Entry(self.frame, width=50)
        self.entry1.grid(row=0, column=1)
        select_btn = tk.Button(self.frame, text="Escoger Estado de Cuenta", command=lambda: self.select_file(self.entry1))
        select_btn.grid(row=0, column=2)

        self.label2 = ttk.Label(self.frame, text="Archivo 2:")
        self.label2.grid(row=1, column=0, sticky=tk.W)
        self.entry2 = ttk.Entry(self.frame, width=50)
        self.entry2.grid(row=1, column=1)
        select_btn2 = tk.Button(self.frame, text="Escoger Auxiliar", command=lambda: self.select_file(self.entry2))
        select_btn2.grid(row=1, column=2)

        self.button_conciliar = ttk.Button(self.frame, text="Conciliar", command=self.conciliar_datos)
        self.button_conciliar.grid(row=2, column=0, columnspan=2)

    def select_file(self, entry):
        file_path = filedialog.askopenfilename(
            title="Select a File",
            initialdir="/",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if file_path:
            entry.delete(0, tk.END)
            entry.insert(0, file_path)
            
    def normalizar_dataframe(self, df):
        # Renombrar columnas para un formato estándar
        columnas = {
            'Descripción': 'Concepto',
            'debe': 'Deposito',
            'haber': 'Retiro'
            }

        df = df.rename(columns=columnas)

        # Asegurar columnas necesarias
        columnas_necesarias = [
            'Fecha',
            'Concepto',
            'Deposito',
            'Retiro'
        ]

        df = df[columnas_necesarias].copy()

        # Limpiar nulos
        df['Deposito'] = df['Deposito'].fillna(0)
        df['Retiro'] = df['Retiro'].fillna(0)

        # Normalizar texto
        df['Concepto'] = (
            df['Concepto']
            .astype(str)
            .str.upper()
            .str.strip()
        )

        # Convertir fecha
        df['Fecha'] = pd.to_datetime(
            df['Fecha'],
            errors='coerce'
        ).dt.date

        return df

    def conciliar_datos(self):
        archivo1 = self.entry1.get()
        archivo2 = self.entry2.get()

        try:

            df1 = pd.read_excel(archivo1)
            df2 = pd.read_excel(archivo2)

            df1 = self.normalizar_dataframe(df1)
            df2 = self.normalizar_dataframe(df2)

            # Separar depósitos
            depositos1 = df1[df1['Deposito'] > 0]
            depositos2 = df2[df2['Deposito'] > 0]
            
            # Separar retiros
            retiros1 = df1[df1['Retiro'] > 0]
            retiros2 = df2[df2['Retiro'] > 0]

            # Ordenar por monto
            depositos1 = depositos1.sort_values(by='Deposito')
            depositos2 = depositos2.sort_values(by='Deposito')
            retiros1 = retiros1.sort_values(by='Retiro')
            retiros2 = retiros2.sort_values(by='Retiro')

            # Listas de resultados
            conciliados = []
            no_conciliados = []

            # Recorrer depósitos
            for _, mov1 in depositos1.iterrows():

                # 1. mismo monto + misma fecha
                coincidencias = depositos2[
                    (depositos2['Deposito'] == mov1['Deposito']) &
                    (depositos2['Fecha'] == mov1['Fecha'])
                ]

                # 2. buscar fecha cercana
                if coincidencias.empty:

                    fecha_inicio = mov1['Fecha'] - pd.Timedelta(days=25)
                    fecha_fin = mov1['Fecha'] + pd.Timedelta(days=25)

                    coincidencias = depositos2[
                        (depositos2['Deposito'] == mov1['Deposito']) &
                        (depositos2['Fecha'] >= fecha_inicio) &
                        (depositos2['Fecha'] <= fecha_fin)
                    ]

                    posibles = pd.DataFrame()
                    # 3. solo monto si hay UNA coincidencia
                    if coincidencias.empty:

                        posibles = depositos2[
                        depositos2['Deposito'] == mov1['Deposito']
                        ]

                    if len(posibles) == 1:
                        coincidencias = posibles
                    
                if not coincidencias.empty:

                    coincidencias = coincidencias.copy()

                    coincidencias['dif_fecha'] = (
                    coincidencias['Fecha'] - mov1['Fecha']
                    ).abs()

                    # elegir la fecha más cercana
                    match = coincidencias.sort_values(
                    by='dif_fecha'
                    ).iloc[0]

                    conciliados.append({
                        'Fecha_1': mov1['Fecha'],
                        'Concepto_1': mov1['Concepto'],
                        'Deposito_1': mov1['Deposito'],

                        'Fecha_2': match['Fecha'],
                        'Concepto_2': match['Concepto'],
                        'Deposito_2': match['Deposito']
                    })

                    depositos2 = depositos2.drop(
                        match.name,
                        errors='ignore'
                    )

                else:
                    no_conciliados.append(mov1)
                    
            #recorrer retiros
            for _, mov1 in retiros1.iterrows():
                coincidencias = retiros2[
                    (retiros2['Retiro'] == mov1['Retiro']) &
                    (retiros2['Fecha'] == mov1['Fecha'])
                ]

                if coincidencias.empty:

                    fecha_inicio = mov1['Fecha'] - pd.Timedelta(days=25)
                    fecha_fin = mov1['Fecha'] + pd.Timedelta(days=25)

                    coincidencias = retiros2[
                        (retiros2['Retiro'] == mov1['Retiro']) &
                        (retiros2['Fecha'] >= fecha_inicio) &
                        (retiros2['Fecha'] <= fecha_fin)
                    ]


                    posibles = pd.DataFrame()
                    
                    if coincidencias.empty:

                        posibles = retiros2[
                        retiros2['Retiro'] == mov1['Retiro']
                        ]

                    if len(posibles) == 1:
                        coincidencias = posibles

                if not coincidencias.empty:

                    coincidencias = coincidencias.copy()

                    coincidencias['dif_fecha'] = (
                    coincidencias['Fecha'] - mov1['Fecha']
                    ).abs()

                    match = coincidencias.sort_values(
                    by='dif_fecha'
                    ).iloc[0]

                    conciliados.append({
                        'Fecha_1': mov1['Fecha'],
                        'Concepto_1': mov1['Concepto'],
                        'Retiro_1': mov1['Retiro'],

                        'Fecha_2': match['Fecha'],
                        'Concepto_2': match['Concepto'],
                        'Retiro_2': match['Retiro']
                    })

                    retiros2 = retiros2.drop(
                        match.name,
                        errors='ignore'
                    )
                    
                else:
                    no_conciliados.append(mov1)
                    
            # convertir listas a DataFrame
            conciliados_df = pd.DataFrame(conciliados)
            no_conciliados_df = pd.DataFrame(no_conciliados)

            # Exportar a Excel
            ruta_salida = filedialog.asksaveasfilename(
                                defaultextension=".xlsx",
                                filetypes=[
                                    ("Excel files", "*.xlsx")
                                ]
                            )

            with pd.ExcelWriter(
                ruta_salida,
                engine='openpyxl'
            ) as writer:

                conciliados_df.to_excel(
                    writer,
                    sheet_name='Conciliados',
                    index=False
                )

                no_conciliados_df.to_excel(
                    writer,
                    sheet_name='No conciliados',
                    index=False
                )

                depositos2.to_excel(
                    writer,
                    sheet_name='No conc aux dep',
                    index=False
                )

                retiros2.to_excel(
                    writer,
                    sheet_name='No conc aux ret',
                    index=False
                )

                # ===== FORMATO =====
                workbook = writer.book

                for sheet_name in writer.sheets:

                    worksheet = writer.sheets[sheet_name]

                    # congelar encabezado
                    worksheet.freeze_panes = 'A2'

                    # activar filtros
                    worksheet.auto_filter.ref = worksheet.dimensions

                    # formato encabezados
                    for cell in worksheet[1]:
                        cell.font = Font(bold=True)

                        cell.fill = PatternFill(
                            start_color="D9EAD3",
                            end_color="D9EAD3",
                            fill_type="solid"
                        )

                    # autoajustar columnas
                    for column_cells in worksheet.columns:

                        length = 0
                        column = column_cells[0].column

                    for cell in column_cells:
                        try:
                            if cell.value:
                                length = max(
                                    length,
                                    len(str(cell.value))
                        )
                        except:
                            pass

                        worksheet.column_dimensions[
                            get_column_letter(column)
                        ].width = length + 2

                    # formato moneda
                    for row in worksheet.iter_rows():

                        for cell in row:

                            if (
                                cell.column_letter in
                                ['C', 'D', 'G', 'H']
                            ):

                                cell.number_format = '#,##0.00'

            print(
                f"Archivo guardado: "
                f"{ruta_salida}"
            )
            
        except Exception as e:
            print(f"Error al leer los archivos: {e}")
            
        messagebox.showinfo("Conciliación Completa", "La conciliación se ha completado. Revisa el archivo de Excel generado.")
            
if __name__ == "__main__":
    root = tk.Tk()
    app = ConciliadorApp(root)
    mainloop()
